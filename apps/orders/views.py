"""Views for checkout and order management."""
import uuid
import razorpay
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.shortcuts import render,redirect, get_object_or_404
from django.urls import reverse
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.db.models import Q
from django.db import transaction
from django.contrib import messages
from apps.userprofile.models import Address
from apps.products.models import Cart,CartItem
from apps.common.decorators import admin_required
from .models import Order, OrderItem, OrderAddress,Payment
from decimal import Decimal
from django.utils import timezone
from django.db import models as db_models
from .models import Coupon, CouponUsage
from .utils import validate_coupon
from apps.products.utils import get_offer_price
from apps.userprofile.wallet_utils import credit_wallet, has_been_refunded
from apps.userprofile.wallet_utils import get_wallet_balance,debit_wallet
from django.db.models import Sum, Count
from django.db.models.functions import TruncDate
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import landscape
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from django.http import HttpResponse
from .utils import release_abandoned_razorpay_orders
from .utils import calculate_item_refund
from apps.userprofile.wallet_utils import has_item_been_refunded


SHIPPING_CHARGES = {
    'standard': 0,
    'express': 25,
}

def generate_order_number():
    return 'ORD' + str(uuid.uuid4().hex[:8]).upper()

def handle_payment(payment_method, order):
    if payment_method == 'cod':
        order.order_status = 'pending'
        order.save()
        return True, "Order Placed Successfully"
    elif payment_method == 'razorpay':
        order.order_status = 'pending'
        order.save()
        return True, "Redirecting to payment gateway..."
    elif payment_method == 'wallet':
        order.order_status = 'processing'
        order.payment_status = 'paid'
        order.save()
        return True, "Order Placed Successfully"
    else:
        return False, "Invalid payment method."

@login_required
@never_cache
def checkout(request):
    user = request.user

    # clean up any abandoned Razorpay orders before showing checkout
    release_abandoned_razorpay_orders(user)

    cart, _ = Cart.objects.get_or_create(user=user)

    cart_items = cart.items.select_related(
        'product_variant',
        'product_variant__product',
        'product_variant__product__category'
    ).all()

    if not cart_items.exists():
        messages.error(request, "Your cart is empty.")
        return redirect('cart')

    removed_items = []

    for item in cart_items:
        variant = item.product_variant
        product = variant.product
        category = product.category

        if (
            variant.status == 'inactive' or
            product.status == 'inactive' or
            category.status == 'inactive'
        ):
            removed_items.append(product.product_name)
            item.delete()

    if removed_items:
        messages.error(
            request,
            f"{', '.join(removed_items)} removed from cart because it is currently unavailable."
        )
        return redirect('cart')

    for item in cart_items:
        variant = item.product_variant
        product = variant.product

        if variant.stock == 0:
            messages.error(
                request,
                f"{product.product_name} is out of stock."
            )
            return redirect('cart')

        if item.quantity > variant.stock:
            messages.error(
                request,
                f"Only {variant.stock} item(s) available for {product.product_name}."
            )
            return redirect('cart')

    addresses = Address.objects.filter(user=user)
    default_address = addresses.filter(is_default=True).first() or addresses.first()

    subtotal = Decimal('0')
    discount_amount = Decimal('0')

    for item in cart_items:
        _, item_discount, _ = get_offer_price(item.product_variant)
        subtotal += item.product_variant.price * item.quantity
        discount_amount += item_discount * item.quantity

    delivery_option = request.GET.get('delivery_option', 'standard')
    shipping_charge = SHIPPING_CHARGES.get(delivery_option, 0)
    total = subtotal + shipping_charge - discount_amount

    coupon_code     = request.session.get('coupon_code')
    coupon_discount = Decimal('0')
    coupon          = None

    if coupon_code:        
        cart_total_for_coupon = subtotal - discount_amount

        revalidated_coupon, revalidated_discount, coupon_error = validate_coupon(
            coupon_code, request.user, cart_total_for_coupon
        )

        if coupon_error:
            # coupon no longer valid
            request.session.pop('coupon_code', None)
            request.session.pop('coupon_discount', None)
            messages.warning(request, f"Your coupon was removed: {coupon_error}")
            coupon = None
            coupon_discount = Decimal('0')
        else:
            coupon = revalidated_coupon
            coupon_discount = revalidated_discount
            # keep session in sync in case the discount amount changed
            request.session['coupon_discount'] = str(coupon_discount)

    final_total = total - coupon_discount

    today = timezone.now().date()
    already_used_ids = CouponUsage.objects.filter(
        user=request.user
    ).values_list('coupon_id', flat=True)

    available_coupons = Coupon.objects.filter(
        status='active',
        expiry_date__gte=today,
    ).exclude(
        id__in=already_used_ids
    ).exclude(
        used_count__gte=db_models.F('usage_limit')
    )

    wallet_balance = get_wallet_balance(request.user)

    context = {
        'addresses': addresses,
        'default_address': default_address,
        'cart': cart,
        'cart_items': cart_items,
        'subtotal': subtotal,
        'discount_amount': discount_amount,
        'shipping_charge': shipping_charge,
        'shipping_label': 'Free' if shipping_charge == 0 else f'₹{shipping_charge}',
        'total': total,
        'delivery_option': delivery_option,
        'coupon': coupon,
        'coupon_discount': coupon_discount,
        'final_total': final_total,
        'available_coupons': available_coupons,
        'wallet_balance': wallet_balance,
    }

    return render(request, 'user/checkout.html', context)

@login_required
@never_cache
def place_order(request):
    if request.method == 'POST':
        user = request.user

        # clean up any abandoned Razorpay orders before placing a new one
        release_abandoned_razorpay_orders(user)

        address_id = request.POST.get('selected_address')
        if not address_id:
            messages.error(request, "Please select a delivery address")
            return redirect('checkout')

        address = get_object_or_404(Address, id=address_id, user=user)

        delivery_option = request.POST.get('delivery_option', 'standard')
        shipping_charge = SHIPPING_CHARGES.get(delivery_option, 0)
        payment_method = request.POST.get('payment_method', 'cod')

        if payment_method == 'wallet':
            pass

        cart = get_object_or_404(Cart, user=user)
        cart_items = cart.items.select_related(
            'product_variant',
            'product_variant__product',
            'product_variant__product__category'
        ).all()

        if not cart_items.exists():
            messages.error(request, "Your cart is empty.")
            return redirect('cart')

        removed_items = []
        for item in cart_items:
            variant = item.product_variant
            product = variant.product
            category = product.category
            if (variant.status == 'inactive' or product.status == 'inactive' or category.status == 'inactive'):
                removed_items.append(product.product_name)
                item.delete()

        if removed_items:
            messages.error(request, f"{', '.join(removed_items)} removed from cart because it is currently unavailable.")
            return redirect('cart')

        cart_items = cart.items.select_related(
            'product_variant', 'product_variant__product', 'product_variant__product__category'
        ).all()

        if not cart_items.exists():
            messages.error(request, "Your cart is empty.")
            return redirect('cart')

        for item in cart_items:
            variant = item.product_variant
            product = variant.product
            if variant.stock == 0:
                messages.error(request, f"{product.product_name} is out of stock.")
                return redirect('cart')
            if item.quantity > variant.stock:
                messages.error(request, f"Only {variant.stock} item(s) available for {product.product_name}.")
                return redirect('cart')
        
        total_amount    = sum(item.product_variant.price * item.quantity for item in cart_items)
        discount_amount = sum(
            (item.product_variant.price * item.quantity) - item.total_price
            for item in cart_items
        )
        final_amount = total_amount + shipping_charge - discount_amount

        if payment_method == 'wallet':
            current_balance = get_wallet_balance(user)
            if current_balance < final_amount:
                messages.error(request, f"Insufficient wallet balance. Available: ₹{current_balance}, Required: ₹{final_amount}.")
                return redirect('checkout')

        try:
            with transaction.atomic():
                order_address = OrderAddress.objects.create(
                    user=user,
                    full_name=address.full_name,
                    phone_number=address.phone_number,
                    address_line1=address.address_line1,
                    address_line2=address.address_line2,
                    city=address.city,
                    state=address.state,
                    country=address.country,
                    pincode=address.pincode,
                    address_type=address.address_type,
                )

                order = Order.objects.create(
                    user=user,
                    order_address=order_address,
                    order_number=generate_order_number(),
                    total_amount=total_amount,
                    discount_amount=discount_amount,
                    final_amount=final_amount,
                    payment_method=payment_method
                )

                # After order is created, handle coupon
                coupon_code = request.session.get('coupon_code')

                if coupon_code:
                    # re-check against the actual cart total being charged right now
                    cart_total_for_coupon = total_amount - discount_amount

                    revalidated_coupon, revalidated_discount, coupon_error = validate_coupon(
                        coupon_code, user, cart_total_for_coupon
                    )

                    if coupon_error:
                        # Coupon no longer valid                        
                        request.session.pop('coupon_code', None)
                        request.session.pop('coupon_discount', None)
                        messages.warning(request, f"Your coupon could not be applied: {coupon_error}")
                    else:
                        order.coupon  = revalidated_coupon
                        order.coupon_discount = revalidated_discount
                        order.final_amount  = order.final_amount - revalidated_discount
                        order.save()

                        CouponUsage.objects.create(
                            coupon=revalidated_coupon,
                            user=user,
                            order=order,
                        )

                        revalidated_coupon.used_count += 1
                        revalidated_coupon.save()

                        if payment_method == 'cod':
                            request.session.pop('coupon_code', None)
                            request.session.pop('coupon_discount', None)

                for item in cart_items:
                    OrderItem.objects.create(
                        order=order,
                        product_variant=item.product_variant,
                        quantity=item.quantity,
                        price=item.product_variant.effective_price,
                        total=item.total_price,
                    )
                    item.product_variant.stock -= item.quantity
                    item.product_variant.save()

                # Create Payment record for both COD and Razorpay
                Payment.objects.create(
                    order=order,
                    amount=order.final_amount,
                    payment_method=payment_method,
                    payment_status='pending',
                )

                if payment_method == 'wallet':
                    debit_wallet(
                        user=user,
                        amount=order.final_amount,
                        description=f"Payment for Order #{order.order_number}",
                        order=order,
                    )
                    order.payment_detail.mark_paid('WALLET-' + order.order_number)

                success, message = handle_payment(payment_method, order)
                if not success:
                    raise Exception(message)

                # Only clear cart for COD — Razorpay clears after payment confirmed
                if payment_method == 'cod':
                    cart_items.delete()

        except Exception as e:
            messages.error(request, str(e))
            return redirect('checkout')
        
        if payment_method == 'razorpay':
            return redirect('initiate_payment', order_id=order.id)
        else:
            messages.success(request, message)
            return redirect('order_confirmation', order_id=order.id)

    return redirect('checkout')

@login_required
@never_cache
def order_confirmation(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    order_items = order.items.select_related('product_variant__product').all()

    context = {
        'order': order,
        'order_items': order_items,
    }
    return render(request, 'user/order_confirmation.html', context)

@admin_required
@never_cache
def order_management(request):
    if not request.user.is_superuser:
        return redirect('admin_login')

    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()
    sort = request.GET.get('sort', '').strip()

    orders = Order.objects.select_related('user', 'order_address').order_by('-created_at')

    if search_query:
        orders = orders.filter(
            Q(order_number__icontains=search_query) |
            Q(user__full_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(order_address__phone_number__icontains=search_query)
        )

    if status_filter:
        orders = orders.filter(order_status=status_filter)

    if sort == 'date_old':
        orders = orders.order_by('created_at')
    elif sort == 'amount_high':
        orders = orders.order_by('-final_amount')
    elif sort == 'amount_low':
        orders = orders.order_by('final_amount')
    else:
        orders = orders.order_by('-created_at')

    paginator = Paginator(orders, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'admin/order_management.html', {
        'orders': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'sort': sort,
        'status_choices': Order.STATUS_CHOICES,
    })

@admin_required
@never_cache
def admin_order_detail(request, order_id):
    order = get_object_or_404(
        Order.objects.select_related('user', 'order_address'),
        id=order_id
    )

    order_items = order.items.select_related('product_variant__product').all()

    # don't show status choices for return_requested
    # it is handled by approve/reject buttons
    if order.order_status == 'return_requested':
        allowed_status_choices = []
    else:
        allowed_values = get_allowed_next_statuses(order.order_status)
        allowed_status_choices = [
            choice for choice in Order.STATUS_CHOICES
            if choice[0] in allowed_values
        ]

    return render(request, 'admin/order_detail.html', {
        'order': order,
        'order_items': order_items,
        'status_choices': allowed_status_choices,
    })

def get_allowed_next_statuses(current_status):
    status_flow = {
        'pending': ['processing', 'cancelled'],
        'processing': ['shipped', 'cancelled'],
        'shipped': ['delivered', 'cancelled'],
        'delivered': ['return_requested'], 
        'return_requested': ['returned', 'delivered'],
        'cancelled': [],
        'returned':  [],
    }

    return status_flow.get(current_status, [])

@admin_required
@never_cache
def update_order_status(request, order_id):
    order = get_object_or_404(Order, id=order_id)

    if request.method == 'POST':

        if order.order_status == 'return_requested':
            messages.error(request, 'Please use Approve or Reject buttons to handle return request.')
            return redirect('admin_order_detail', order_id=order.id)

        new_status = request.POST.get('order_status')
        new_payment_status = request.POST.get('payment_status')

        allowed_statuses = get_allowed_next_statuses(order.order_status)

        if new_status not in allowed_statuses:
            messages.error(request, 'Invalid status transition.')
            return redirect('admin_order_detail', order_id=order.id)

        if new_status == 'cancelled':
            total_refunded = Decimal('0')

            for item in order.items.filter(status='active').select_related('product_variant'):
                item.product_variant.stock += item.quantity
                item.product_variant.save()
                item.status = 'cancelled'
                item.save()

                if order.payment_status == 'paid':
                    refund_amount = calculate_item_refund(item)
                    if refund_amount > 0 and not has_item_been_refunded(item):
                        credit_wallet(
                            user=order.user,
                            amount=refund_amount,
                            description=f"Refund for cancelled item ({item.product_variant}) in Order #{order.order_number}",
                            order=order,
                            order_item=item,
                        )
                        total_refunded += refund_amount

            order.order_status = new_status   # ← THE MISSING LINE

            if total_refunded > 0:
                messages.success(request, f"Order cancelled. ₹{total_refunded} refunded to customer's wallet.")
            else:
                messages.success(request, 'Order cancelled successfully.')
        else:
            order.order_status = new_status   # ← THE MISSING LINE (for every other transition too)
            messages.success(request, 'Order status updated successfully.')

        if new_payment_status in ['pending', 'paid']:
            order.payment_status = new_payment_status

        order.save()

    return redirect('admin_order_detail', order_id=order.id)


@admin_required
@never_cache
def update_item_status(request, order_id, item_id):
    if not request.user.is_superuser:
        return redirect('admin_login')

    item = get_object_or_404(OrderItem, id=item_id, order_id=order_id)
    order = item.order

    if request.method == 'POST':
        new_status = request.POST.get('item_status')

        if new_status not in ['active', 'cancelled']:
            messages.error(request, 'Invalid item status.')
            return redirect('admin_order_detail', order_id=order_id)

        if new_status == 'cancelled' and item.status != 'cancelled':
            item.product_variant.stock += item.quantity
            item.product_variant.save()

            if order.payment_status == 'paid':
                refund_amount = calculate_item_refund(item)
                if refund_amount > 0 and not has_item_been_refunded(item):
                    credit_wallet(
                        user=order.user,
                        amount=refund_amount,
                        description=f"Refund for cancelled item ({item.product_variant}) in Order #{order.order_number}",
                        order=order,
                        order_item=item,
                    )

            item.status = 'cancelled'
            item.save()

            all_cancelled = not order.items.filter(status='active').exists()
            if all_cancelled:
                order.order_status = 'cancelled'
                order.save()

            messages.success(request, 'Item cancelled, stock restored, and refund processed if applicable.')
        else:
            item.status = new_status
            item.save()
            messages.success(request, f'Item status updated to {new_status}.')

    return redirect('admin_order_detail', order_id=order_id)

@admin_required
@never_cache
def handle_return_request(request, order_id):
    order = get_object_or_404(Order, id=order_id)

    if order.order_status != 'return_requested':
        messages.error(request, "This order has no pending return request.")
        return redirect('admin_order_detail', order_id=order.id)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'approve':
            # restore stock
            for item in order.items.select_related('product_variant').all():
                item.product_variant.stock += item.quantity
                item.product_variant.save()

            order.order_status = 'returned'
            order.save()
             # Wallet refund after admin approval
            if order.payment_status == 'paid' and not has_been_refunded(order):
                credit_wallet(
                    user=order.user,
                    amount=order.final_amount,
                    description=f"Refund for returned Order #{order.order_number}",
                    order=order,
                )
                messages.success(request, f"Return approved. Stock restored and ₹{order.final_amount} refunded to customer's wallet.")
            else:
                messages.success(request, "Return approved. Stock has been restored.")


        elif action == 'reject':
            # put order back to delivered
            order.order_status = 'delivered'
            order.return_reason = None
            order.save()
            messages.success(request, "Return request rejected. Order is back to delivered.")

        else:
            messages.error(request, "Invalid action.")

    return redirect('admin_order_detail', order_id=order.id)

@admin_required
@never_cache
def handle_item_return_request(request, order_id, item_id):
    order = get_object_or_404(Order, id=order_id)
    item = get_object_or_404(OrderItem, id=item_id, order=order)

    if item.status != 'return_requested':
        messages.error(request, "This item has no pending return request.")
        return redirect('admin_order_detail', order_id=order.id)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'approve':
            item.product_variant.stock += item.quantity
            item.product_variant.save()

            item.status = 'returned'
            item.save()

            refund_amount = calculate_item_refund(item)

            if order.payment_status == 'paid' and refund_amount > 0 and not has_item_been_refunded(item):
                credit_wallet(
                    user=order.user,
                    amount=refund_amount,
                    description=f"Refund for returned item ({item.product_variant}) in Order #{order.order_number}",
                    order=order,
                    order_item=item,
                )
                messages.success(request, f"Item return approved. Stock restored and ₹{refund_amount} refunded to customer's wallet.")
            else:
                messages.success(request, "Item return approved. Stock has been restored.")

            if not order.items.filter(status__in=['active', 'return_requested']).exists():
                order.order_status = 'returned'
                order.save()

        elif action == 'reject':
            item.status = 'active'
            item.return_reason = None
            item.save()
            messages.success(request, "Item return request rejected.")

        else:
            messages.error(request, "Invalid action.")

    return redirect('admin_order_detail', order_id=order.id)

razorpay_client = razorpay.Client(
    auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET)
)

# Initiate Payment
@login_required
def initiate_payment(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)

    # Amount in paise (₹1 = 100 paise)
    amount_paise = int(order.final_amount * 100)

    # Create order on Razorpay's server
    rzp_order = razorpay_client.order.create({
        "amount":          amount_paise,
        "currency":        "INR",
        "payment_capture": 1,
        "notes":           {"order_number": order.order_number},
    })

    # Save/update Payment record
    payment, _ = Payment.objects.get_or_create(
        order=order,
        defaults={'amount': order.final_amount}
    )
    payment.razorpay_order_id = rzp_order['id']
    payment.save()

    return render(request, "user/razorpay_checkout.html", {
        "order":             order,
        "razorpay_order_id": rzp_order['id'],
        "razorpay_key_id":   settings.RAZORPAY_KEY_ID,
        "amount_paise":      amount_paise,
        "user_name":         request.user.full_name,
        "user_email":        request.user.email,
    })


# Verify Payment (called after Razorpay popup)
@csrf_exempt
def verify_payment(request):
    if request.method != "POST":
        return redirect('home')

    rzp_order_id   = request.POST.get('razorpay_order_id')
    rzp_payment_id = request.POST.get('razorpay_payment_id')
    rzp_signature  = request.POST.get('razorpay_signature')
    order_id       = request.POST.get('order_id')

    order   = get_object_or_404(Order, id=order_id)
    payment = get_object_or_404(Payment, order=order)

    try:
        # Verify signature — prevents fake payments
        razorpay_client.utility.verify_payment_signature({
            'razorpay_order_id':   rzp_order_id,
            'razorpay_payment_id': rzp_payment_id,
            'razorpay_signature':  rzp_signature,
        })

        # Mark payment and order as paid
        payment.mark_paid(rzp_payment_id)
        order.payment_status = 'paid'
        order.order_status   = 'processing'
        order.save()

        # Clear coupon from session now that payment is confirmed
        request.session.pop('coupon_code', None)
        request.session.pop('coupon_discount', None)
        
        try:
            cart = Cart.objects.get(user=order.user)
            CartItem.objects.filter(cart=cart).delete()
        except Cart.DoesNotExist:
            pass

        return redirect('payment_success', order_id=order.id)
    
    except razorpay.errors.SignatureVerificationError:
        payment.mark_failed()
        order.order_status = 'cancelled'
        order.save()

        # Restore stock
        for item in order.items.select_related('product_variant').all():
            item.product_variant.stock += item.quantity
            item.product_variant.save()

        # rollback coupon usage
        if order.coupon:
            CouponUsage.objects.filter(
                coupon=order.coupon,
                user=order.user
            ).delete()
            order.coupon.used_count = max(0, order.coupon.used_count - 1)
            order.coupon.save()
            # Keep coupon in session so user sees it on checkout

        return redirect(f"{reverse('payment_failure', args=[order.id])}?reason=verification_failed")


# Success Page
@login_required
def payment_success(request, order_id):
    order   = get_object_or_404(Order, id=order_id, user=request.user)
    payment = get_object_or_404(Payment, order=order)
    return render(request, "user/payment_success.html", {
        "order":   order,
        "payment": payment,
    })


# Failure Page 
@login_required
def payment_failure(request, order_id):
    order   = get_object_or_404(Order, id=order_id, user=request.user)
    payment = get_object_or_404(Payment, order=order)
    reason  = request.GET.get('reason', 'unknown')

    # If user cancelled (dismissed popup) and stock not yet restored
    if reason == 'cancelled' and order.order_status == 'pending':
        order.order_status = 'cancelled'
        order.save()
        payment.mark_failed()

        # Restore stock
        for item in order.items.select_related('product_variant').all():
            item.product_variant.stock += item.quantity
            item.product_variant.save()

        # rollback coupon usage so user can use it again
        if order.coupon:
            # Remove usage record
            CouponUsage.objects.filter(
                coupon=order.coupon,
                user=request.user
            ).delete()

            # Decrement used count
            order.coupon.used_count = max(0, order.coupon.used_count - 1)
            order.coupon.save()

            # Coupon stays in session so it shows on checkout again

    return render(request, "user/payment_failure.html", {
        "order":  order,
        "reason": reason,
    })


@login_required
def apply_coupon(request):
    if request.method != 'POST':
        return redirect('checkout')

    code = request.POST.get('coupon_code', '').strip().upper()

    # Prevent applying if coupon already applied
    if request.session.get('coupon_code'):
        messages.error(request, "A coupon is already applied. Remove it first.")
        return redirect('checkout')

    # Get current cart total (after offer discounts)
    try:
        cart  = Cart.objects.get(user=request.user)
        items = CartItem.objects.filter(cart=cart).select_related(
            'product_variant__product__category'
        )
        cart_total = sum(
            get_offer_price(item.product_variant)[0] * item.quantity
            for item in items
        )
    except Cart.DoesNotExist:
        messages.error(request, "Your cart is empty.")
        return redirect('checkout')

    coupon, discount, error = validate_coupon(code, request.user, cart_total)

    if error:
        messages.error(request, error)
        return redirect('checkout')

    # Store in session
    request.session['coupon_code']     = coupon.coupon_code
    request.session['coupon_discount'] = str(discount)

    messages.success(request, f"Coupon '{coupon.coupon_code}' applied! You save ₹{discount}.")
    return redirect('checkout')


@login_required
def remove_coupon(request):
    if request.method != 'POST':
        return redirect('checkout')

    request.session.pop('coupon_code', None)
    request.session.pop('coupon_discount', None)

    messages.success(request, "Coupon removed successfully.")
    return redirect('checkout')

#Coupon Management (Admin)
@admin_required
@never_cache
def coupon_management(request):
    search_query = request.GET.get('search', '').strip()

    coupons = Coupon.objects.order_by('-created_at')

    if search_query:
        coupons = coupons.filter(
            Q(coupon_code__icontains=search_query) |
            Q(discount_type__icontains=search_query)
        )

    paginator = Paginator(coupons, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'admin/coupon_management.html', {
        'coupons': page_obj,
        'search_query': search_query,
    })

@admin_required
@never_cache
def add_coupon(request):
    if request.method == 'POST':
        code            = request.POST.get('coupon_code', '').strip().upper()
        discount_type   = request.POST.get('discount_type', '').strip()
        discount_value  = request.POST.get('discount_value', '').strip()
        minimum_price   = request.POST.get('minimum_price', '0').strip()
        maximum_redeem  = request.POST.get('maximum_redeem', '').strip()
        expiry_date     = request.POST.get('expiry_date', '').strip()
        usage_limit     = request.POST.get('usage_limit', '1').strip() 

        # Coupon code
        if not code:
            messages.error(request, "Coupon code is required.")
            return render(request, 'admin/add_coupon.html', status=400)

        if len(code) < 3:
            messages.error(request, "Coupon code must be at least 3 characters.")
            return render(request, 'admin/add_coupon.html', status=400)

        if len(code) > 20:
            messages.error(request, "Coupon code cannot exceed 20 characters.")
            return render(request, 'admin/add_coupon.html', status=400)

        if not code.isalnum():
            messages.error(request, "Coupon code can only contain letters and numbers.")
            return render(request, 'admin/add_coupon.html', status=400)

        if Coupon.objects.filter(coupon_code=code).exists():
            messages.error(request, "Coupon code already exists.")
            return render(request, 'admin/add_coupon.html', status=400)

        # Discount type
        if discount_type not in ['percentage', 'flat']:
            messages.error(request, "Please select a valid discount type.")
            return render(request, 'admin/add_coupon.html', status=400)

        # Discount value
        if not discount_value:
            messages.error(request, "Discount value is required.")
            return render(request, 'admin/add_coupon.html', status=400)

        try:
            discount_value_num = float(discount_value)
        except ValueError:
            messages.error(request, "Discount value must be a valid number.")
            return render(request, 'admin/add_coupon.html', status=400)

        if discount_value_num <= 0:
            messages.error(request, "Discount value must be greater than 0.")
            return render(request, 'admin/add_coupon.html', status=400)

        if discount_type == 'percentage' and discount_value_num > 100:
            messages.error(request, "Percentage discount cannot exceed 100.")
            return render(request, 'admin/add_coupon.html', status=400)

        if discount_type == 'flat' and discount_value_num > 100000:
            messages.error(request, "Flat discount amount seems unreasonably high.")
            return render(request, 'admin/add_coupon.html', status=400)

        # Minimum order amount
        try:
            minimum_price_num = float(minimum_price) if minimum_price else 0
        except ValueError:
            messages.error(request, "Minimum order amount must be a valid number.")
            return render(request, 'admin/add_coupon.html', status=400)

        if minimum_price_num < 0:
            messages.error(request, "Minimum order amount cannot be negative.")
            return render(request, 'admin/add_coupon.html', status=400)

        if discount_type == 'flat' and minimum_price_num <= discount_value_num:
            messages.error(request, "Minimum order amount must be greater than the flat discount value.")
            return render(request, 'admin/add_coupon.html', status=400)

        # Maximum redeem
        maximum_redeem_num = None
        if maximum_redeem:
            try:
                maximum_redeem_num = float(maximum_redeem)
            except ValueError:
                messages.error(request, "Maximum redeem amount must be a valid number.")
                return render(request, 'admin/add_coupon.html', status=400)

            if maximum_redeem_num <= 0:
                messages.error(request, "Maximum redeem amount must be greater than 0.")
                return render(request, 'admin/add_coupon.html', status=400)

        # Expiry date
        if not expiry_date:
            messages.error(request, "Expiry date is required.")
            return render(request, 'admin/add_coupon.html', status=400)

        try:
            expiry_date_obj = datetime.strptime(expiry_date, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid expiry date format.")
            return render(request, 'admin/add_coupon.html', status=400)

        today = timezone.now().date()
        if expiry_date_obj <= today:
            messages.error(request, "Expiry date must be in the future.")
            return render(request, 'admin/add_coupon.html', status=400)

        # Usage limit
        if not usage_limit:
            messages.error(request, "Usage limit is required.")
            return render(request, 'admin/add_coupon.html', status=400)

        try:
            usage_limit_num = int(usage_limit)
        except ValueError:
            messages.error(request, "Usage limit must be a whole number.")
            return render(request, 'admin/add_coupon.html', status=400)

        if usage_limit_num < 1:
            messages.error(request, "Usage limit must be at least 1.")
            return render(request, 'admin/add_coupon.html', status=400)

        Coupon.objects.create(
            coupon_code    = code,
            discount_type  = discount_type,
            discount_value = discount_value_num,
            minimum_price  = minimum_price_num,
            maximum_redeem = maximum_redeem_num,
            expiry_date    = expiry_date_obj,
            usage_limit    = usage_limit_num,
        )
        messages.success(request, "Coupon created successfully!")
        return redirect('coupon_management')

    return render(request, 'admin/add_coupon.html')

@admin_required
@never_cache
def edit_coupon(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)

    if request.method == 'POST':
        discount_value = request.POST.get('discount_value', '').strip()
        minimum_price = request.POST.get('minimum_price', '0').strip()
        maximum_redeem = request.POST.get('maximum_redeem', '').strip()
        expiry_date  = request.POST.get('expiry_date', '').strip()
        usage_limit  = request.POST.get('usage_limit', '1').strip()

        if not discount_value:
            messages.error(request, "Discount value is required.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        try:
            discount_value_num = float(discount_value)
        except ValueError:
            messages.error(request, "Discount value must be a valid number.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        if discount_value_num <= 0:
            messages.error(request, "Discount value must be greater than 0.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        if coupon.discount_type == 'percentage' and discount_value_num > 100:
            messages.error(request, "Percentage discount cannot exceed 100.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        try:
            minimum_price_num = float(minimum_price) if minimum_price else 0
        except ValueError:
            messages.error(request, "Minimum order amount must be a valid number.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        if minimum_price_num < 0:
            messages.error(request, "Minimum order amount cannot be negative.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)
        
        if coupon.discount_type == 'flat' and minimum_price_num <= discount_value_num:
            messages.error(request, "Minimum order amount must be greater than the flat discount value.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        maximum_redeem_num = None
        if maximum_redeem:
            try:
                maximum_redeem_num = float(maximum_redeem)
            except ValueError:
                messages.error(request, "Maximum redeem amount must be a valid number.")
                return redirect('edit_coupon', coupon_id=coupon.id, status=400)

            if maximum_redeem_num <= 0:
                messages.error(request, "Maximum redeem amount must be greater than 0.")
                return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        if not expiry_date:
            messages.error(request, "Expiry date is required.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        try:
            expiry_date_obj = datetime.strptime(expiry_date, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid expiry date format.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        if not usage_limit:
            messages.error(request, "Usage limit is required.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        try:
            usage_limit_num = int(usage_limit)
        except ValueError:
            messages.error(request, "Usage limit must be a whole number.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        if usage_limit_num < 1:
            messages.error(request, "Usage limit must be at least 1.")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        if usage_limit_num < coupon.used_count:
            messages.error(request, f"Usage limit cannot be less than the number of times already used ({coupon.used_count}).")
            return redirect('edit_coupon', coupon_id=coupon.id, status=400)

        coupon.discount_value  = discount_value_num
        coupon.minimum_price   = minimum_price_num
        coupon.maximum_redeem  = maximum_redeem_num
        coupon.expiry_date     = expiry_date_obj
        coupon.usage_limit     = usage_limit_num
        coupon.save()
        messages.success(request, "Coupon updated successfully!")
        return redirect('coupon_management')

    return render(request, 'admin/edit_coupon.html', {'coupon': coupon})

@admin_required
@never_cache
def toggle_coupon_status(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    coupon.status = 'inactive' if coupon.status == 'active' else 'active'
    coupon.save()
    messages.success(request, f"Coupon {'activated' if coupon.status == 'active' else 'deactivated'} successfully.")
    return redirect('coupon_management')


@admin_required
@never_cache
def delete_coupon(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    coupon.delete()
    messages.success(request, "Coupon deleted successfully.")
    return redirect('coupon_management')


def get_date_range(request):
    range_type = request.GET.get('range', 'monthly')
    today = timezone.now().date()

    if range_type == 'daily':
        start = today
        end = today
        label = 'Today'
    elif range_type == 'weekly':
        start = today - timedelta(days=today.weekday())
        end = today
        label = 'This Week'
    elif range_type == 'yearly':
        start = today.replace(month=1, day=1)
        end = today
        label = 'This Year'
    elif range_type == 'custom':
        start_str = request.GET.get('start_date')
        end_str = request.GET.get('end_date')
        try:
            start = datetime.strptime(start_str, '%Y-%m-%d').date()
            end = datetime.strptime(end_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            start = today.replace(day=1)
            end = today
        label = f'{start.strftime("%b %d, %Y")} - {end.strftime("%b %d, %Y")}'
    else:
        range_type = 'monthly'
        start = today.replace(day=1)
        end = today
        label = 'This Month'

    return start, end, range_type, label


def get_sales_queryset(start, end):
    return Order.objects.filter(
        created_at__date__gte=start,
        created_at__date__lte=end,
    ).exclude(order_status='cancelled')


@admin_required
@never_cache
def sales_report(request):
    start, end, range_type, label = get_date_range(request)
    orders = get_sales_queryset(start, end)

    summary = orders.aggregate(
        total_orders=Count('id'),
        total_revenue=Sum('final_amount'),
        total_discount=Sum('discount_amount'),
        total_coupon_discount=Sum('coupon_discount'),
    )

    total_orders = summary['total_orders'] or 0
    total_revenue = summary['total_revenue'] or Decimal('0')
    total_discount = (summary['total_discount'] or Decimal('0')) + (summary['total_coupon_discount'] or Decimal('0'))

    daily_breakdown = (
        orders
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(
            order_count=Count('id'),
            revenue=Sum('final_amount'),
            discount=Sum('discount_amount'),
            coupon_discount=Sum('coupon_discount'),
        )
        .order_by('-day')
    )

    paginator = Paginator(list(daily_breakdown), 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'range_type': range_type,
        'range_label': label,
        'start_date': start,
        'end_date': end,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'total_discount': total_discount,
        'daily_breakdown': page_obj,
    }
    return render(request, 'admin/sales_report.html', context)


@admin_required
@never_cache
def sales_report_pdf(request):
    start, end, _, label = get_date_range(request)
    orders = get_sales_queryset(start, end)

    summary = orders.aggregate(
        total_orders=Count('id'),
        total_revenue=Sum('final_amount'),
        total_discount=Sum('discount_amount'),
        total_coupon_discount=Sum('coupon_discount'),
    )
    total_orders = summary['total_orders'] or 0
    total_revenue = summary['total_revenue'] or Decimal('0')
    total_discount = (summary['total_discount'] or Decimal('0')) + (summary['total_coupon_discount'] or Decimal('0'))

    daily_breakdown = (
        orders
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(
            order_count=Count('id'),
            revenue=Sum('final_amount'),
            discount=Sum('discount_amount'),
            coupon_discount=Sum('coupon_discount'),
        )
        .order_by('-day')
    )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{start}_{end}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('title', fontSize=20, spaceAfter=6, fontName='Helvetica-Bold')
    elements.append(Paragraph("SCENTORA - Sales Report", title_style))
    elements.append(Paragraph(f"Period: {label} ({start} to {end})", styles['Normal']))
    elements.append(Spacer(1, 0.2 * inch))

    summary_data = [
        ['Total Orders', 'Total Revenue', 'Total Discount'],
        [str(total_orders), f"Rs.{total_revenue}", f"Rs.{total_discount}"],
    ]
    summary_table = Table(summary_data, colWidths=[3*inch, 3*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 14),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3 * inch))

    elements.append(Paragraph("Daily Breakdown", styles['Heading3']))
    table_data = [['Date', 'Orders', 'Revenue', 'Discount', 'Coupon Discount']]
    for row in daily_breakdown:
        table_data.append([
            row['day'].strftime('%b %d, %Y') if row['day'] else '-',
            str(row['order_count']),
            f"Rs.{row['revenue'] or 0}",
            f"Rs.{row['discount'] or 0}",
            f"Rs.{row['coupon_discount'] or 0}",
        ])

    detail_table = Table(table_data, colWidths=[2*inch, 1.5*inch, 2*inch, 2*inch, 2*inch])
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(detail_table)

    doc.build(elements)
    return response


@admin_required
@never_cache
def sales_report_excel(request):
    start, end, range_type, label = get_date_range(request)
    orders = get_sales_queryset(start, end)

    daily_breakdown = (
        orders
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(
            order_count=Count('id'),
            revenue=Sum('final_amount'),
            discount=Sum('discount_amount'),
            coupon_discount=Sum('coupon_discount'),
        )
        .order_by('-day')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    ws.merge_cells('A1:E1')
    ws['A1'] = f"SCENTORA Sales Report - {label} ({start} to {end})"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Date', 'Orders', 'Revenue (Rs)', 'Discount (Rs)', 'Coupon Discount (Rs)']
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row in daily_breakdown:
        ws.append([
            row['day'].strftime('%Y-%m-%d') if row['day'] else '-',
            row['order_count'],
            float(row['revenue'] or 0),
            float(row['discount'] or 0),
            float(row['coupon_discount'] or 0),
        ])

    for col_letter, width in zip('ABCDE', [15, 10, 15, 15, 18]):
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sales_report_{start}_{end}.xlsx"'
    wb.save(response)
    return response